import requests
import openpyxl as xl
from bs4 import BeautifulSoup
import time
from requests import TooManyRedirects

'''
checking 240 webPages for 120s

inputs [ name of xl or url of webPage ::  Error tages => attribute and other ] 
create xl file with url pages


path of xl => str
number of sheet (1,2,3,4...) => int
position of cell (A,1) => int [2]
take teg name => str
attributes => None or str [2]

'''


# web = 'https://google.ru/fdsa'
# r = requests.get(web)
# if r.status_code != 200:
#     print(r.status_code)
#     exit()
# else:
#     r = requests.get(web).text
#
# soup = BeautifulSoup(r, "lxml")
# block = soup.find("title").text
# if block != 'Error 404 (Not Found)!!1':
#     print(f'accept')
# else:
#     print(f'Error -- {web}')
# exit()


# wb = xl.load_workbook('webPages/allWeb.xlsx')
# sheet = wb['Sheet2']
# ws = wb.active
#
# for i in range(2, sheet.max_row + 1):
#     # time.sleep(3)
#     cell = sheet.cell(i, 2)
#     web = cell.value
#     web = web.replace('new.', '')  # comment this string
#     try:
#         r = requests.get(web)
#     except TooManyRedirects:
#         print(f'Error TooManyRedirects {i}')
#         continue
#     r = requests.get(web).text
#     soup = BeautifulSoup(r, "lxml")
#     block = soup.find("section", class_="error-wrapper")
#     if block is None:
#         print(f'accept {i}')
#     else:
#         print(f'Error {i} -- {web}')


def web_pages(path, numberSheet, positionFirst, positionSecond, tagName):
    wb = xl.load_workbook(f'{path}')
    sheet = wb[f'{numberSheet}']
    ws = wb.active

    for i in range(positionFirst, sheet.max_row + 1):
        # time.sleep(3)
        cell = sheet.cell(i, positionSecond)
        web = cell.value
        web = web.replace('new.', '')  # comment this string
        try:
            r = requests.get(web)
        except TooManyRedirects:
             print(f'Error TooManyRedirects {i}')
             continue
        r = requests.get(web).text
        soup = BeautifulSoup(r, "lxml")
        block = soup.find(f"{tagName}", class_=f"{atributClass}", id=f"{atributId}")
        if block is None:
            print(f'accept {i}')
        else:
            print(f'Error {i} -- {web}')


path = str(input('Путь xl файла (пр. webPages/allWeb.xlsx):'))
numberSheet = str(input('который из Sheet-ов (пр. Sheet2):'))
positionFirst = int(input('место ячейи(A=1,B=2,C=3...):'))
positionSecond = int(input('место ячейи(1,2,3...):'))
tagName = str(input('названия тега (пр. section):'))
atributClass = str(input('названия тега (пр. error-wrapper):'))
atributId = str(input('названия тега (пр. main-content):'))

web_pages(path, numberSheet, positionFirst, positionSecond, tagName)

#           inputs
# webPages/allWeb.xlsx
# Sheet2
# 2
# 2
# section
